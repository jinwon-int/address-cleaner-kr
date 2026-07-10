"""cli.py 서브커맨드 왕복 테스트 (네트워크 없이)."""

from __future__ import annotations

import json

import openpyxl

from address_cleaner.cli import main
from address_cleaner.clients import SearchResult


def _workbook(tmp_path, rows: list[str]):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    for i, addr in enumerate(rows, start=2):
        ws[f"H{i}"] = addr
    wb.save(input_path)
    return input_path


# --- normalize ---


def test_normalize_prints_json_schema(capsys):
    exit_code = main(
        ["normalize", "경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호"]
    )

    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert exit_code == 0
    assert payload["query"] == "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    assert payload["kind"] == "lot"
    assert payload["status"] == "ok"
    assert set(payload) >= {"original", "query", "kind", "status", "detail"}


# --- excel ---


def test_excel_roundtrip_with_provider_none(tmp_path, capsys):
    input_path = _workbook(tmp_path, ["경기도 파주시 야당동 57-17", "주소 미상"])
    output_path = tmp_path / "out.xlsx"

    exit_code = main(
        [
            "excel",
            str(input_path),
            "-o",
            str(output_path),
            "--status-col",
            "M",
            "--provider",
            "none",
        ]
    )

    captured = capsys.readouterr()
    stats = json.loads(captured.out)
    assert exit_code == 0
    assert stats["total"] == 2
    assert stats["missing"] == 1
    ws = openpyxl.load_workbook(output_path).active
    assert ws["I2"].value == "경기도 파주시 야당동 57-17"
    assert ws["M3"].value == "검색주소없음"


def test_excel_returns_2_for_missing_input_file(tmp_path, capsys):
    exit_code = main(
        ["excel", str(tmp_path / "없는파일.xlsx"), "-o", str(tmp_path / "out.xlsx")]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")


def test_excel_detail_col_without_status_col_fails_cleanly(tmp_path, capsys):
    input_path = _workbook(tmp_path, ["경기도 파주시 야당동 57-17"])

    exit_code = main(
        [
            "excel",
            str(input_path),
            "-o",
            str(tmp_path / "out.xlsx"),
            "--detail-col",
            "N",
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--status-col" in captured.err


# --- feedback ---


def _failure_workbook(tmp_path):
    input_path = tmp_path / "result.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"], ws["I1"], ws["M1"], ws["N1"] = (
        "원주소",
        "주소검색어",
        "처리결과",
        "처리상세",
    )
    ws["H2"] = "경기도 파주시 야당동 57-17"
    ws["I2"] = "경기도 파주시 야당동 57-17"
    ws["M2"] = "완료"
    ws["H3"] = "경기도 파주시 야당동 57-17 정우펠리스 제1(상층하층)층 제101호"
    ws["I3"] = "경기도 파주시 야당동 57-17 정우펠리스 제1(상층하층)층 제101호"
    ws["M3"] = "실패(우편번호미확정)"
    ws["N3"] = "정보요청팝업[우편번호미확정]"
    wb.save(input_path)
    return input_path


def test_feedback_prints_report_to_stdout(tmp_path, capsys):
    input_path = _failure_workbook(tmp_path)

    exit_code = main(["feedback", str(input_path)])

    captured = capsys.readouterr()
    report = json.loads(captured.out)
    assert exit_code == 0
    assert report["failures"] == 1
    assert report["byResult"] == {"실패(우편번호미확정)": 1}
    assert report["rows"][0]["row"] == 3


def test_feedback_saves_report_to_file(tmp_path, capsys):
    input_path = _failure_workbook(tmp_path)
    report_path = tmp_path / "feedback.json"

    exit_code = main(["feedback", str(input_path), "-o", str(report_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert str(report_path) in captured.out
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["failures"] == 1


def test_feedback_returns_2_for_missing_input_file(tmp_path, capsys):
    exit_code = main(["feedback", str(tmp_path / "없는파일.xlsx")])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")


# --- probe ---


def test_probe_juso_prints_result(monkeypatch, capsys):
    class _FakeJuso:
        def __init__(self):
            self.key = "test-key"

        def search(self, query: str, count: int = 10):
            return SearchResult("juso", 1, {"roadAddr": "경기도 파주시 하우3길 22"}, {})

    monkeypatch.setattr("address_cleaner.cli.JusoClient", _FakeJuso)

    exit_code = main(["probe", "juso", "경기도 파주시 야당동 57-17"])

    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert exit_code == 0
    assert payload["provider"] == "juso"
    assert payload["found"] is True
    assert payload["total_count"] == 1


def test_probe_epost_falls_back_to_compact_query(monkeypatch, capsys):
    class _FakeEpost:
        def __init__(self):
            self.key = "test-key"
            self.queries: list[str] = []

        def search(self, query: str, search_se: str = "road", count: int = 10):
            self.queries.append(query)
            if query == "하우3길 22":
                return SearchResult("epost", 1, {"zipNo": "10911"}, "")
            return SearchResult("epost", 0, {}, "")

    monkeypatch.setattr("address_cleaner.cli.KoreaPostRoadNameClient", _FakeEpost)

    exit_code = main(["probe", "epost", "경기도 파주시 하우3길 22 정우펠리스 제101호"])

    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert exit_code == 0
    assert payload["provider"] == "epost"
    assert payload["query"] == "하우3길 22"  # 실제 사용한 검색어를 알려준다
    assert payload["found"] is True
