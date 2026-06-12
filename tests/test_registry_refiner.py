from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

from address_cleaner.registry.excel import refine
from address_cleaner.registry.juso import first_pass_status, make_queries
from address_cleaner.registry.normalize import (
    clean_raw,
    has_extra_parcels,
    load_typo_rules,
    original_is_under_specified,
    parse_lot_addr,
    set_extra_typo_rules,
    strip_detail,
    strip_unit,
    typo_fix,
)


FIXTURES = Path(__file__).parent / "fixtures" / "anonymized_addresses.json"


class AddressRefinementTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.fixtures = json.loads(FIXTURES.read_text(encoding="utf-8"))

    def test_clean_and_parse_anonymized_lot_addresses(self) -> None:
        for case in self.fixtures:
            with self.subTest(case=case["name"]):
                cleaned = clean_raw(case["raw"])

                self.assertEqual(cleaned, case["expected_clean"])
                self.assertEqual(strip_unit(cleaned), case["expected_unitless"])
                self.assertEqual(parse_lot_addr(cleaned)["addr"], case["expected_lot_addr"])

    def test_make_queries_keeps_lot_lookup_candidate(self) -> None:
        for case in self.fixtures:
            with self.subTest(case=case["name"]):
                queries = make_queries(case["raw"], "", "", "")

                self.assertIn(case["expected_lot_addr"], queries)
                self.assertLessEqual(len(queries), 12)


class TypoFixTest(unittest.TestCase):
    def test_keeps_road_names_starting_with_si(self) -> None:
        self.assertEqual(typo_fix("인천광역시 시민로 100"), "인천광역시 시민로 100")

    def test_collapses_duplicated_si_token(self) -> None:
        self.assertEqual(typo_fix("인천광역시 시 부평구 부평동 100-1"), "인천광역시 부평구 부평동 100-1")

    def test_expands_sido_abbreviation_only_at_start(self) -> None:
        self.assertEqual(typo_fix("서울 강남구 역삼동 123-4"), "서울특별시 강남구 역삼동 123-4")
        self.assertEqual(
            typo_fix("인천광역시 연수구 옥련동 12-3 서울 빌라 101호"),
            "인천광역시 연수구 옥련동 12-3 서울 빌라 101호",
        )

    def test_strips_leading_zip_before_expanding_abbreviation(self) -> None:
        self.assertEqual(typo_fix("12345 경기 김포시 사우동 256-1"), "경기도 김포시 사우동 256-1")


class ExtraTypoRulesTest(unittest.TestCase):
    def tearDown(self) -> None:
        set_extra_typo_rules(None)

    def test_extra_rules_apply_after_loading(self) -> None:
        self.assertEqual(typo_fix("서울특별시 강남구 역삼동 12-3 헬리오시티이"), "서울특별시 강남구 역삼동 12-3 헬리오시티이")

        set_extra_typo_rules([("헬리오시티이", "헬리오시티")])

        self.assertEqual(typo_fix("서울특별시 강남구 역삼동 12-3 헬리오시티이"), "서울특별시 강남구 역삼동 12-3 헬리오시티")

    def test_load_typo_rules_accepts_list_and_dict_forms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            list_path = Path(tmp) / "rules_list.json"
            list_path.write_text('[["가나다", "가나도"]]', encoding="utf-8")
            dict_path = Path(tmp) / "rules_dict.json"
            dict_path.write_text('{"replacements": [["가나다", "가나도"]]}', encoding="utf-8")

            self.assertEqual(load_typo_rules(list_path), [("가나다", "가나도")])
            self.assertEqual(load_typo_rules(dict_path), [("가나다", "가나도")])

    def test_load_typo_rules_rejects_malformed_entries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            bad_path = Path(tmp) / "rules_bad.json"
            bad_path.write_text('[["하나만"]]', encoding="utf-8")

            with self.assertRaises(ValueError):
                load_typo_rules(bad_path)


class ParseLotAddrTest(unittest.TestCase):
    def test_mountain_lot_number(self) -> None:
        parsed = parse_lot_addr("서울특별시 관악구 신림동 산 101-1 가나빌라 201호")

        self.assertEqual(parsed["lot"], "산101-1")
        self.assertEqual(parsed["addr"], "서울특별시 관악구 신림동 산101-1")

    def test_sido_omitted_address(self) -> None:
        self.assertEqual(parse_lot_addr("수원시 팔달구 인계동 123-4")["addr"], "수원시 팔달구 인계동 123-4")

    def test_unit_number_is_not_a_lot(self) -> None:
        self.assertEqual(parse_lot_addr("서울특별시 성동구 마장동 801호")["addr"], "")

    def test_dong_without_district_is_not_an_address(self) -> None:
        self.assertEqual(parse_lot_addr("마장동 801")["addr"], "")


class StripDetailTest(unittest.TestCase):
    def test_keeps_lot_only_address(self) -> None:
        self.assertEqual(strip_detail("서울특별시 강남구 역삼동 123"), "서울특별시 강남구 역삼동 123")

    def test_strips_bare_trailing_unit_number(self) -> None:
        self.assertEqual(strip_detail("서울특별시 강남구 역삼동 123-4 1202"), "서울특별시 강남구 역삼동 123-4")

    def test_keeps_road_building_number(self) -> None:
        self.assertEqual(strip_detail("서울특별시 강남구 테헤란로 123 1202"), "서울특별시 강남구 테헤란로 123")


class ExtraParcelsTest(unittest.TestCase):
    def test_detects_extra_parcels(self) -> None:
        self.assertTrue(has_extra_parcels("인천광역시 연수구 옥련동 123-4 외 3필지 101호"))

    def test_ignores_oe_inside_building_name(self) -> None:
        self.assertFalse(has_extra_parcels("인천광역시 연수구 옥련동 123-4 외동마을 101호"))


class UnderSpecifiedTest(unittest.TestCase):
    def test_dong_and_unit_only_is_under_specified(self) -> None:
        self.assertTrue(original_is_under_specified("서울특별시 성동구 마장동 801호"))

    def test_mountain_lot_is_specified(self) -> None:
        self.assertFalse(original_is_under_specified("서울특별시 관악구 신림동 산 101-1 201호"))

    def test_building_name_is_specified(self) -> None:
        self.assertFalse(original_is_under_specified("서울특별시 성동구 마장동 가람빌리지 801호"))




class RegistryExactOnePolicyTest(unittest.TestCase):
    @staticmethod
    def _juso_row(unit: str = "101") -> dict[str, str]:
        return {
            "roadAddr": f"서울특별시 강남구 테헤란로 123 테스트빌딩 {unit}호",
            "roadAddrPart1": "서울특별시 강남구 테헤란로 123",
            "jibunAddr": "서울특별시 강남구 역삼동 123-4",
            "bdNm": "테스트빌딩",
            "zipNo": "06234",
            "admCd": "1168010100",
            "rnMgtSn": "116803122001",
            "udrtYn": "0",
            "buldMnnm": "123",
            "buldSlno": "0",
            "bdMgtSn": f"BD-{unit}",
            "lnbrMnnm": "123",
            "lnbrSlno": "4",
        }

    def _refine_one_row_with_first_pass(self, first_total: int) -> dict[str, str | int | None]:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_dir = tmp_path / "out"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "대상 임대차계약 주소"
            ws["A2"] = "서울특별시 강남구 역삼동 123-4 테스트빌딩 101호"
            wb.save(input_file)

            row = self._juso_row("101")
            call_count = {"first": 0}

            def fake_juso_query(_session, _key, keyword, _cache, count=5, preserve_commas=False):
                if preserve_commas:
                    call_count["first"] += 1
                    if call_count["first"] == 1:
                        rows = [row.copy() for _ in range(first_total)]
                        return {"keyword": keyword, "total": first_total, "rows": rows[:count]}
                    return {"keyword": keyword, "total": 0, "rows": []}
                return {"keyword": keyword, "total": 1, "rows": [row]}

            with mock.patch("address_cleaner.registry.excel.juso_query", side_effect=fake_juso_query):
                refine(input_file, output_dir, "dummy-key", verbose=False)

            output_file = output_dir / "input_등기소전체검색용_최종검색주소추가.xlsx"
            out_wb = openpyxl.load_workbook(output_file)
            out_ws = out_wb["Sheet1"]
            headers = {out_ws.cell(1, c).value: c for c in range(1, out_ws.max_column + 1)}
            return {
                "JUSO_판정": out_ws.cell(2, headers["JUSO_판정"]).value,
                "JUSO_총건수": out_ws.cell(2, headers["JUSO_총건수"]).value,
                "JUSO_2차판정": out_ws.cell(2, headers["JUSO_2차판정"]).value,
                "주소검토결과": out_ws.cell(2, headers["주소검토결과"]).value,
                "등기_검토사유": out_ws.cell(2, headers["등기_검토사유"]).value,
            }

    def test_first_pass_marks_zero_results_as_not_found(self) -> None:
        status, best = first_pass_status({"total": 0, "keyword": "원문", "rows": []}, None)

        self.assertEqual(status, "검색불가")
        self.assertEqual(best["total"], 0)

    def test_first_pass_marks_multiple_original_results_as_ambiguous(self) -> None:
        status, best = first_pass_status({"total": 2, "keyword": "원문", "rows": [{}, {}]}, None)

        self.assertEqual(status, "다중검출_원문")
        self.assertEqual(best["total"], 2)

    def test_first_pass_marks_multiple_normalized_results_as_ambiguous(self) -> None:
        status, best = first_pass_status(
            {"total": 0, "keyword": "원문", "rows": []},
            {"total": 3, "keyword": "상세제거", "rows": [{}, {}, {}]},
        )

        self.assertEqual(status, "다중검출_상세주소제거")
        self.assertEqual(best["total"], 3)

    def test_zero_result_first_pass_stays_review_even_after_high_second_pass(self) -> None:
        row = self._refine_one_row_with_first_pass(0)

        self.assertEqual(row["JUSO_판정"], "검색불가")
        self.assertEqual(row["JUSO_2차판정"], "자동추천_높음")
        self.assertEqual(row["주소검토결과"], "검토후조회")
        self.assertIn("Juso 1차 검색불가", row["등기_검토사유"] or "")

    def test_multiple_result_first_pass_stays_review_even_after_high_second_pass(self) -> None:
        row = self._refine_one_row_with_first_pass(2)

        self.assertEqual(row["JUSO_판정"], "다중검출_원문")
        self.assertEqual(row["JUSO_2차판정"], "자동추천_높음")
        self.assertEqual(row["주소검토결과"], "검토후조회")
        self.assertIn("Juso 1차 다중검출_원문", row["등기_검토사유"] or "")

class BackwardCompatImportTest(unittest.TestCase):
    def test_cli_still_re_exports_refactored_functions(self) -> None:
        from address_cleaner.registry.cli import (  # noqa: F401
            clean_raw,
            make_queries,
            parse_lot_addr,
            refine,
            strip_unit,
            typo_fix,
        )


class DuplexFloorNotationTest(unittest.TestCase):
    def test_clean_raw_handles_duplex_floor_with_parenthetical(self):
        # 실제 전산 입력 실패 사례(우편번호 미확정)에서 확인된 복층 표기.
        self.assertEqual(
            clean_raw("경기도 고양시 일산동구 중산동 78-7 시크릿타운 제비동 제1(상층하층)층 제101호"),
            "경기도 고양시 일산동구 중산동 78-7 시크릿타운 비동 1층 101호",
        )


if __name__ == "__main__":
    unittest.main()
