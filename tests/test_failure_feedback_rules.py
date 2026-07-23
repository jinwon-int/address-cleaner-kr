"""실전 실패 사례(우편번호미확정) 환류로 추가된 정제 규칙 테스트.

각 케이스는 파워쉘 주소찾기 팝업에서 '실패(우편번호미확정)'로 끝난 실제
원주소 패턴을 익명화 수준만 유지한 채 그대로 옮긴 것이다.
"""

import openpyxl

from address_cleaner.clients import SearchResult
from address_cleaner.excel import _verify, process_workbook
from address_cleaner.normalizer import base_for_search, normalize_for_search


def test_road_bungil_number_stays_glued_to_road_name():
    # '권광로175번길'을 '권광로175 번길'로 가르면 도로명이 깨져 검색이 실패한다.
    result = normalize_for_search(
        "경기 수원시 팔달구 권광로175번길91 (인계동,ECO HYGGE 1)--408"
    )
    assert result.kind == "road"
    assert result.query.startswith("경기 수원시 팔달구 권광로175번길 91")
    assert "--" not in result.query
    assert result.query.endswith("408호")


def test_numbered_dong_ga_stays_one_token():
    # '양평동4가'는 서수 법정동 한 토큰이다. '양평동 4가'로 가르면 안 된다.
    result = normalize_for_search(
        "서울 영등포구 양평동4가 1-9 어반52 501호 (도로명 : 선유로44길 3)"
    )
    assert result.kind == "lot"
    assert result.query == "서울 영등포구 양평동4가 1-9 어반52 501호"
    assert base_for_search(result.query, result.kind) == "서울 영등포구 양평동4가 1-9"


def test_et_al_household_marker_is_removed():
    result = normalize_for_search(
        "서울 강서구 화곡동 1169 더숨153 105동 203호 외 3세대"
    )
    assert result.query == "서울 강서구 화곡동 1169 더숨153 105동 203호"


def test_legal_area_description_is_cut_and_unit_recovered():
    result = normalize_for_search(
        "경기 용인시 기흥구 하갈동 297-23외 1필지 옥스빌 3층 151.45㎡ 중 도면표시 "
        "ㄱ,ㄴ,ㄷ,ㄹ,ㅁ,ㅂ,ㅅ,ㄱ의 각 점을 순차로 연결한 선 내 부분 302호 전체 67.74㎡"
    )
    assert result.query == "경기 용인시 기흥구 하갈동 297-23 옥스빌 302호"
    assert result.kind == "lot"


def test_legal_drawing_description_without_unit_leaves_skeleton():
    result = normalize_for_search(
        "경기도 수원시 권선구 세류동 218-100외 1필지 2층 194.49㎡ 중 도면표시 "
        "가,나,다,라,가의 각 점을 순차로 연결한 선내부분 24㎡"
    )
    assert result.query == "경기도 수원시 권선구 세류동 218-100"


def test_dev_district_block_lot_tokens_are_removed():
    result = normalize_for_search(
        "인천 남동구 논현동 751-1번지 인천소래논현구역씨10블럭 에코메트로3차더타워 A동 1802호"
    )
    assert result.query == "인천 남동구 논현동 751-1 에코메트로3차더타워 A동 1802호"


def test_dev_district_prefix_glued_to_building_is_removed():
    result = normalize_for_search(
        "인천 남동구 논현동 751-1 소래논현구역씨10블록에코메트로3차더타워 제씨동 제4503호"
    )
    assert (
        result.query == "인천 남동구 논현동 751-1 에코메트로3차더타워 제씨동 제4503호"
    )


def test_apartment_complex_name_with_sindosi_is_not_split():
    result = normalize_for_search(
        "경기도 화성시 동탄구 장지동 977 동탄2신도시호반베르디움33단지 제3315동 제6층 제601호"
    )
    assert (
        result.query
        == "경기도 화성시 동탄구 장지동 977 동탄2신도시호반베르디움33단지 제3315동 제601호"
    )


def test_dong_echo_paren_is_removed():
    result = normalize_for_search("인천 미추홀구 주안동 295-20 1104 (주안동)")
    assert result.query == "인천 미추홀구 주안동 295-20 1104"


def test_dash_dong_orphan_is_removed():
    cases = [
        ("서울 강서구 양천로 714-14-동 902호", "서울 강서구 양천로 714-14 902호"),
        ("서울 강서구 양천로 714-14 -동 802호", "서울 강서구 양천로 714-14 802호"),
    ]
    for raw, expected in cases:
        result = normalize_for_search(raw)
        assert result.query == expected
        assert result.kind == "road"


def test_letter_dash_unit_becomes_identified_dong():
    result = normalize_for_search("서울 강서구 마곡동 793 제비-508호")
    assert result.query == "서울 강서구 마곡동 793 비동 508호"


def test_basement_floor_and_dong_typo_are_normalized():
    # '제2등'은 '제2동' 오타, '제지하층'은 층 정보라 제거된다.
    result = normalize_for_search(
        "인천 미추홀구 용현동 456-320 보은빌라 제2등 제지하층 제비02호"
    )
    assert result.query == "인천 미추홀구 용현동 456-320 보은빌라 제2동 제비02호"


def test_letter_glued_floor_keeps_dong_and_drops_floor():
    result = normalize_for_search(
        "서울 강서구 화곡동 372-141 신영베르디움 비제2층 제204호"
    )
    assert result.query == "서울 강서구 화곡동 372-141 신영베르디움 비동 제204호"


def test_officetel_word_glued_to_unit_is_removed():
    result = normalize_for_search(
        "인천 미추홀구 주안동 77-3 주안지웰어스테이트 102동 오피스텔410호"
    )
    assert result.query == "인천 미추홀구 주안동 77-3 주안지웰어스테이트 102동 410호"


def test_building_name_ending_in_officetel_is_kept():
    result = normalize_for_search("경기 김포시 풍무동 196 대우오피스텔 205호")
    assert result.query == "경기 김포시 풍무동 196 대우오피스텔 205호"


def test_duplicated_unit_phrases_are_collapsed():
    result = normalize_for_search(
        "인천 미추홀구 숭의동 352-21 제비동 402호 (숭의동, 한마음아파트) 제비동동402호"
    )
    assert result.query == "인천 미추홀구 숭의동 352-21 제비동 402호 한마음아파트"


def test_bare_unit_then_identified_unit_keeps_identified_form():
    result = normalize_for_search(
        "서울 영등포구 버드나루로 9, 604호 (영등포동2가) 1동 604호"
    )
    assert result.query == "서울 영등포구 버드나루로 9 1동 604호"


def test_vowel_jamo_residue_after_digit_is_removed():
    result = normalize_for_search("인천 남동구 논현동 751-1 제1ㅣ동 제1305호")
    assert result.query == "인천 남동구 논현동 751-1 제1동 제1305호"


def test_sido_typo_gwangyeosit_is_corrected():
    result = normalize_for_search(
        "인천광여깃 남동구 구월동 1183-15 수도팬더빌라 제1층 제101호"
    )
    assert result.query == "인천광역시 남동구 구월동 1183-15 수도팬더빌라 제101호"


# ---- 건물명 검색어 (kind="building") ----


def test_building_query_without_lot_number():
    result = normalize_for_search("서울 강북구 수유동 한양아이빌 비동 402호")
    assert result.kind == "building"
    assert result.searchable
    assert result.query == "서울 강북구 수유동 한양아이빌 비동 402호"
    assert result.detail == "비동 402호"


def test_building_name_is_adopted_from_annotation_paren():
    result = normalize_for_search("서울 강서구 화곡동 B동 503호 (화곡동, 타운캐슬)")
    assert result.kind == "building"
    assert result.query == "서울 강서구 화곡동 타운캐슬 B동 503호"


def test_building_name_glued_to_dong_after_sigungu_is_split():
    result = normalize_for_search("인천광역시 미추홀구 주안동정다운파크빌 306호")
    assert result.kind == "building"
    assert result.query == "인천광역시 미추홀구 주안동 정다운파크빌 306호"


def test_building_query_keeps_dev_district_out(  # 지구/블럭/로트 잡음 제거와의 결합
):
    result = normalize_for_search(
        "경기 김포시 고촌읍 신곡리 김포신곡6지구 도시개발사업구역 에이2블럭 "
        "캐슬앤파밀리에시티2단지 213동 405호"
    )
    assert result.kind == "building"
    assert (
        result.query == "경기 김포시 고촌읍 신곡리 캐슬앤파밀리에시티2단지 213동 405호"
    )


def test_unit_only_addresses_stay_unsearchable():
    for raw in [
        "인천 남동구 간석동 503호",
        "경기 고양시 일산동구 풍동 102동 202호",
        "인천 부평구 부평동 5층 504호",
        "경기 김포시 통진읍 303호",
    ]:
        result = normalize_for_search(raw)
        assert result.query == ""
        assert result.kind == "invalid"
        assert result.status == "unrecognized"


def test_base_for_search_building_is_district_and_name():
    assert (
        base_for_search("서울 강북구 수유동 한양아이빌 비동 402호", "building")
        == "서울 강북구 수유동 한양아이빌"
    )


def test_verify_building_falls_back_to_district_and_name_base():
    class _BaseOnlyJuso:
        key = "test-key"

        def __init__(self):
            self.queries = []

        def search(self, query: str, count: int = 5):
            self.queries.append(query)
            if query == "서울 강북구 수유동 한양아이빌":
                return SearchResult(
                    "juso",
                    1,
                    {"roadAddr": "서울특별시 강북구 삼양로 100", "zipNo": "01000"},
                    {},
                )
            return SearchResult("juso", 0, {}, {})

    juso = _BaseOnlyJuso()
    status, detail, correction = _verify(
        "서울 강북구 수유동 한양아이빌 비동 402호", "building", juso, None
    )

    assert status == "verified"
    assert juso.queries == [
        "서울 강북구 수유동 한양아이빌 비동 402호",
        "서울 강북구 수유동 한양아이빌",
    ]
    assert correction["type"] == "상세부제거"
    assert correction["working"] == "서울 강북구 수유동 한양아이빌"


# ---- 검증 확정 골격으로 검색어 교체 ----


def _base_only_workbook(tmp_path):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 고양시 일산동구 중산동 78-7 시크릿타운 제비동 제101호"
    wb.save(input_path)
    return input_path


class _BaseOnlySkeletonJuso:
    """상세 포함 검색은 0건, 골격 검색만 1건을 돌려주는 가짜 Juso."""

    key = "test-key"

    def search(self, query: str, count: int = 5):
        if query == "경기도 고양시 일산동구 중산동 78-7":
            return SearchResult(
                "juso",
                1,
                {"roadAddr": "경기도 고양시 일산동구 중앙로 123", "zipNo": "10401"},
                {},
            )
        return SearchResult("juso", 0, {}, {})


def test_process_workbook_rewrites_query_with_verified_skeleton(tmp_path, monkeypatch):
    monkeypatch.setattr("address_cleaner.excel.JusoClient", _BaseOnlySkeletonJuso)
    input_path = _base_only_workbook(tmp_path)
    output_path = tmp_path / "out.xlsx"

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        detail_col="N",
        provider="juso",
        mark_missing=True,
    )

    result_ws = openpyxl.load_workbook(output_path).active
    # 상세 포함 검색어는 팝업에서 0건이 확정이므로, 통한 골격으로 교체된다.
    assert result_ws["I2"].value == "경기도 고양시 일산동구 중산동 78-7"
    assert result_ws["M2"].value in (None, "")
    assert "골격으로 교체" in result_ws["N2"].value
    assert stats["query_rewritten"] == 1
    assert stats["verified"] == 1


def test_process_workbook_keeps_detail_query_when_disabled(tmp_path, monkeypatch):
    monkeypatch.setattr("address_cleaner.excel.JusoClient", _BaseOnlySkeletonJuso)
    input_path = _base_only_workbook(tmp_path)
    output_path = tmp_path / "out.xlsx"

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        provider="juso",
        mark_missing=True,
        rewrite_working_query=False,
    )

    result_ws = openpyxl.load_workbook(output_path).active
    assert (
        result_ws["I2"].value
        == "경기도 고양시 일산동구 중산동 78-7 시크릿타운 제비동 제101호"
    )
    assert stats["query_rewritten"] == 0
