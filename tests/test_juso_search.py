"""juso_search 공용 인프라와 excel 병렬 검증 테스트."""

from __future__ import annotations

import threading
import time

import openpyxl
import pytest

from address_cleaner.clients import SearchResult
from address_cleaner.excel import STATUS_AMBIGUOUS, STATUS_NOT_FOUND, process_workbook
from address_cleaner.juso_search import RateLimiter


# --- RateLimiter ---


def test_rate_limiter_enforces_calls_per_second_floor():
    limiter = RateLimiter(50.0)  # 20ms 간격
    calls = 6

    def hit():
        limiter.wait()

    threads = [threading.Thread(target=hit) for _ in range(calls)]
    start = time.monotonic()
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    elapsed = time.monotonic() - start

    # 첫 호출은 즉시, 나머지 5회는 20ms 간격 → 최소 0.1초 소요
    assert elapsed >= (calls - 1) * (1.0 / 50.0)


def test_rate_limiter_zero_rate_never_blocks():
    limiter = RateLimiter(0)
    start = time.monotonic()
    for _ in range(100):
        limiter.wait()
    assert time.monotonic() - start < 0.1


def test_registry_import_paths_stay_compatible():
    # 기존 자동화가 쓰는 등기 모드 경로가 그대로 동작해야 한다.
    from address_cleaner import juso_search
    from address_cleaner.registry import juso as registry_juso

    assert registry_juso.RateLimiter is juso_search.RateLimiter
    assert registry_juso.juso_query is juso_search.juso_query
    assert registry_juso.load_cache is juso_search.load_cache
    assert registry_juso.save_cache is juso_search.save_cache
    assert registry_juso.cache_lock is juso_search.cache_lock
    assert registry_juso.set_rate_limiter is juso_search.set_rate_limiter


# --- 직렬 vs 병렬 동등성 ---

ADDRESSES = {
    # 상세포함 1건 → verified
    "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호": 1,
    # 2건 이상 → ambiguous
    "서울특별시 강남구 테헤란로 152 강남파이낸스센터": 2,
    # 0건 (골격/변형도 0건) → missing
    "경기도 파주시 야당동 99-99 없는빌라 101호": 0,
}


class _MappedJuso:
    """검색어 → 건수 맵으로 동작하는 결정적 가짜 클라이언트 (스레드 안전 기록)."""

    key = "test-key"

    calls: list[str] = []
    _lock = threading.Lock()

    def search(self, query: str, count: int = 5):
        with self._lock:
            _MappedJuso.calls.append(query)
        for addr, total in ADDRESSES.items():
            if query == addr or addr.startswith(query):
                return SearchResult(
                    "juso", total, {"roadAddr": "표준주소", "zipNo": "10911"}, {}
                )
        return SearchResult("juso", 0, {}, {})


def _build_workbook(tmp_path, name):
    input_path = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    rows = list(ADDRESSES) + [
        list(ADDRESSES)[0],  # 중복 행: 병렬 제출 전에 dedup 되어야 한다
        "주소 미상",  # 로컬 불가
    ]
    for i, addr in enumerate(rows, start=2):
        ws[f"H{i}"] = addr
    wb.save(input_path)
    return input_path


@pytest.mark.parametrize("workers", [1, 8])
def test_serial_and_parallel_produce_identical_output(tmp_path, monkeypatch, workers):
    monkeypatch.setattr("address_cleaner.excel.JusoClient", _MappedJuso)
    _MappedJuso.calls = []
    input_path = _build_workbook(tmp_path, f"input_{workers}.xlsx")
    output_path = tmp_path / f"out_{workers}.xlsx"

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        detail_col="N",
        provider="juso",
        mark_missing=True,
        workers=workers,
    )

    assert stats["verified"] == 2  # 중복 행 포함 2행 verified
    assert stats["ambiguous"] == 1
    assert stats["missing"] == 2  # API 0건 1행 + 로컬 불가 1행
    ws = openpyxl.load_workbook(output_path).active
    assert ws["M2"].value in (None, "")
    assert ws["M3"].value == STATUS_AMBIGUOUS
    assert ws["M4"].value == STATUS_NOT_FOUND
    assert ws["M5"].value in (None, "")  # 중복 행도 같은 판정
    assert ws["M6"].value == STATUS_NOT_FOUND
    assert "1건" in ws["N2"].value
    # 중복 주소는 병렬 제출 전에 제거된다: 상세포함 1건 주소는 딱 1회 검색
    assert _MappedJuso.calls.count(list(ADDRESSES)[0]) == 1


def test_parallel_and_serial_cells_match(tmp_path, monkeypatch):
    monkeypatch.setattr("address_cleaner.excel.JusoClient", _MappedJuso)
    outputs = {}
    all_stats = {}
    for workers in (1, 8):
        input_path = _build_workbook(tmp_path, f"in_{workers}.xlsx")
        output_path = tmp_path / f"cmp_{workers}.xlsx"
        all_stats[workers] = process_workbook(
            input_path,
            output_path,
            source_col="H",
            target_col="I",
            status_col="M",
            detail_col="N",
            provider="juso",
            mark_missing=True,
            workers=workers,
        )
        ws = openpyxl.load_workbook(output_path).active
        outputs[workers] = [
            (ws[f"I{r}"].value, ws[f"M{r}"].value, ws[f"N{r}"].value)
            for r in range(2, ws.max_row + 1)
        ]

    assert outputs[1] == outputs[8]
    assert all_stats[1] == all_stats[8]


def test_parallel_worker_runtime_error_propagates(tmp_path, monkeypatch):
    class _DeadKeyJuso:
        key = "expired-key"

        def search(self, query: str, count: int = 5):
            return SearchResult(
                "juso",
                0,
                {"errorCode": "E0014", "errorMessage": "승인되지 않은 KEY"},
                {},
            )

    monkeypatch.setattr("address_cleaner.excel.JusoClient", _DeadKeyJuso)
    input_path = _build_workbook(tmp_path, "dead.xlsx")

    with pytest.raises(RuntimeError, match="API errors"):
        process_workbook(
            input_path,
            tmp_path / "out.xlsx",
            source_col="H",
            target_col="I",
            status_col="M",
            provider="juso",
            mark_missing=True,
            workers=8,
        )


def test_excel_cli_accepts_workers(tmp_path, capsys):
    input_path = _build_workbook(tmp_path, "cli.xlsx")
    from address_cleaner.cli import main

    exit_code = main(
        [
            "excel",
            str(input_path),
            "-o",
            str(tmp_path / "out.xlsx"),
            "--provider",
            "none",
            "--workers",
            "4",
        ]
    )

    capsys.readouterr()
    assert exit_code == 0
