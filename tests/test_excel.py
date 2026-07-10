"""excel.py 테스트: 열 변환, 워크북 라운드트립 규약, 옵션 검증.

_verify 판정 분기와 이력/교정 리포트는 test_normalizer.py에서 이미 다루므로
여기서는 그 밖의 규약(열 인덱스, 옵션 조합, 라운드트립 형식)을 굳힌다.
"""

from __future__ import annotations

import openpyxl
import pytest

from address_cleaner.excel import col_to_index, process_workbook


# --- col_to_index ---


@pytest.mark.parametrize(
    ("col", "index"),
    [("A", 1), ("H", 8), ("Z", 26), ("AA", 27), ("AZ", 52), ("h", 8), (" m ", 13)],
)
def test_col_to_index_supports_multi_letter_and_lowercase(col, index):
    assert col_to_index(col) == index


@pytest.mark.parametrize("col", ["", "1", "A1", "가", "-"])
def test_col_to_index_rejects_non_alpha(col):
    with pytest.raises(ValueError):
        col_to_index(col)


# --- 옵션 조합 ---


def test_detail_col_requires_status_col(tmp_path):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    wb.active["H1"] = "원주소"
    wb.save(input_path)

    with pytest.raises(RuntimeError, match="--detail-col requires --status-col"):
        process_workbook(
            input_path,
            tmp_path / "out.xlsx",
            source_col="H",
            target_col="I",
            detail_col="N",
        )


# --- provider=none 라운드트립 규약 ---


def test_provider_none_roundtrip_writes_headers_and_queries(tmp_path):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호"
    ws["H3"] = "주소 미상"
    ws["A4"] = "원주소 없는 서식 행"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        tmp_path / "out.xlsx",
        source_col="H",
        target_col="I",
        status_col="M",
        provider="none",
    )

    ws = openpyxl.load_workbook(tmp_path / "out.xlsx").active
    assert ws["I1"].value == "주소검색어"
    assert ws["M1"].value == "주소검색결과"
    # 정상 행: I열 검색어, M열 빈칸
    assert ws["I2"].value == "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    assert ws["M2"].value in (None, "")
    # 무효 행: 검색어 비우고 검색주소없음
    assert not ws["I3"].value
    assert ws["M3"].value == "검색주소없음"
    # 빈 행: M열 공란 (후단 자동화가 불량 주소와 혼동하지 않도록)
    assert ws["M4"].value in (None, "")
    assert stats == {
        **stats,
        "total": 3,
        "lot": 1,
        "invalid": 1,
        "empty": 1,
        "missing": 1,
    }


def test_header_false_processes_first_row(tmp_path):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    wb.active["H1"] = "경기도 파주시 야당동 57-17"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        tmp_path / "out.xlsx",
        source_col="H",
        target_col="I",
        provider="none",
        header=False,
    )

    ws = openpyxl.load_workbook(tmp_path / "out.xlsx").active
    assert ws["I1"].value == "경기도 파주시 야당동 57-17"
    assert stats["total"] == 1
