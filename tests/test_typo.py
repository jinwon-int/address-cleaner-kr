"""typo.py 공용 오타 규칙 테스트: 일반 정제기 반영과 CLI --typo-rules 왕복."""

from __future__ import annotations

import json

import openpyxl
import pytest

from address_cleaner.cli import main
from address_cleaner.normalizer import normalize_for_search
from address_cleaner.typo import (
    apply_typo_replacements,
    load_typo_rules,
    set_extra_typo_rules,
)


@pytest.fixture(autouse=True)
def _reset_extra_rules():
    # 전역 규칙 상태가 다른 테스트로 새지 않게 한다 (기존 registry 테스트 패턴 동일).
    yield
    set_extra_typo_rules(None)


def test_base_rules_fix_sido_typo_in_normalizer():
    result = normalize_for_search("서울틀벽시 강남구 역삼동 123-4 어떤빌라 101호")

    assert result.kind == "lot"
    assert result.query.startswith("서울특별시 강남구 역삼동 123-4")


def test_extra_rule_applies_to_normalize_for_search():
    raw = "경기도 파주시 야당동 57-17 정우팰리스 101호"
    before = normalize_for_search(raw).query
    assert "정우팰리스" in before  # 기본 규칙만으로 교정된다면 이 테스트 전제가 깨진 것

    set_extra_typo_rules([("정우팰리스", "정우펠리스")])
    after = normalize_for_search(raw).query

    assert "정우펠리스" in after
    assert "정우팰리스" not in after


def test_without_extra_rules_result_is_unchanged():
    raw = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호"
    baseline = normalize_for_search(raw)

    set_extra_typo_rules([("없는패턴", "다른패턴")])
    with_noop_rule = normalize_for_search(raw)

    assert with_noop_rule == baseline


def test_apply_typo_replacements_is_pure_replacement():
    # norm()과 달리 쉼표/괄호를 건드리지 않는다 — normalizer의 자체 규칙에 맡긴다.
    assert (
        apply_typo_replacements("서울시 강남구 (역삼동, 어떤빌라)")
        == "서울특별시 강남구 (역삼동, 어떤빌라)"
    )


def test_load_typo_rules_accepts_list_and_dict_forms(tmp_path):
    list_path = tmp_path / "rules_list.json"
    list_path.write_text(
        json.dumps([["프루지오", "푸르지오"]], ensure_ascii=False), encoding="utf-8"
    )
    dict_path = tmp_path / "rules_dict.json"
    dict_path.write_text(
        json.dumps({"replacements": [["게양대로", "계양대로"]]}, ensure_ascii=False),
        encoding="utf-8",
    )

    assert load_typo_rules(list_path) == [("프루지오", "푸르지오")]
    assert load_typo_rules(dict_path) == [("게양대로", "계양대로")]


def test_registry_import_paths_stay_compatible():
    # 기존 자동화가 쓰는 등기 모드 경로가 그대로 동작해야 한다.
    from address_cleaner.registry import load_typo_rules as registry_load
    from address_cleaner.registry.normalize import (
        BASE_TYPO_REPLACEMENTS,
        typo_fix,
    )

    assert registry_load is load_typo_rules
    assert ("프루지오", "푸르지오") in BASE_TYPO_REPLACEMENTS
    assert typo_fix("서울 강남구 역삼동 123-4") == "서울특별시 강남구 역삼동 123-4"


def test_excel_cli_applies_typo_rules(tmp_path, capsys):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps([["정우팰리스", "정우펠리스"]], ensure_ascii=False),
        encoding="utf-8",
    )
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우팰리스 101호"
    wb.save(input_path)
    output_path = tmp_path / "out.xlsx"

    exit_code = main(
        [
            "excel",
            str(input_path),
            "-o",
            str(output_path),
            "--provider",
            "none",
            "--typo-rules",
            str(rules_path),
        ]
    )

    capsys.readouterr()
    assert exit_code == 0
    ws = openpyxl.load_workbook(output_path).active
    assert ws["I2"].value == "경기도 파주시 야당동 57-17 정우펠리스 101호"


def test_excel_cli_reports_malformed_typo_rules(tmp_path, capsys):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text('[["규칙만하나"]]', encoding="utf-8")
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    wb.active["H1"] = "원주소"
    wb.save(input_path)

    exit_code = main(
        [
            "excel",
            str(input_path),
            "-o",
            str(tmp_path / "out.xlsx"),
            "--typo-rules",
            str(rules_path),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
