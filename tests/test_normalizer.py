from pathlib import Path

import openpyxl
import requests

from address_cleaner.clients import SearchResult
from address_cleaner.cli import main
from address_cleaner.excel import STATUS_AMBIGUOUS, STATUS_NOT_FOUND, _verify, process_workbook
from address_cleaner.normalizer import compact_for_epost, normalize_for_search, preprocess_raw_address


def test_lot_address_keeps_building_detail():
    result = normalize_for_search("경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호")
    assert result.query == "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    assert result.kind == "lot"
    assert result.searchable


def test_lot_address_removes_extra_lot_marker():
    result = normalize_for_search("경기도 파주시 동패동 7외 1필지 노블타운 제102동 제4층 제401호")
    assert result.query == "경기도 파주시 동패동 7 노블타운 제102동 제401호"
    assert result.kind == "lot"


def test_road_address_keeps_detail_after_building_number():
    result = normalize_for_search("서울특별시 강남구 테헤란로 152 강남파이낸스센터 10층")
    assert result.query == "서울특별시 강남구 테헤란로 152 강남파이낸스센터"
    assert result.kind == "road"


def test_lot_detail_excludes_spaced_mountain_lot_number():
    result = normalize_for_search("서울특별시 강남구 역삼동 산 12-3 어떤빌라 101호")
    assert result.kind == "lot"
    assert result.detail == "어떤빌라 101호"


def test_compact_for_epost_uses_road_name_and_building_number_only():
    assert compact_for_epost("경기도 파주시 하우3길 22 정우펠리스 제303동 제101호", "road") == "하우3길 22"


def test_compact_for_epost_uses_lot_area_and_lot_number_only():
    assert compact_for_epost("경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호", "lot") == "야당동 57-17"


def test_preprocess_repairs_missing_spaces():
    assert preprocess_raw_address("경기도파주시야당동57-17") == "경기도 파주시 야당동 57-17"


def test_malformed_address_is_not_searchable():
    result = normalize_for_search("주소 미상")
    assert result.query == ""
    assert result.kind == "invalid"
    assert not result.searchable


def test_unrecognized_address_does_not_emit_fallback_query():
    result = normalize_for_search("경기도 파주시 정우펠리스 제303동 제1층 제101호")
    assert result.query == ""
    assert result.kind == "invalid"
    assert result.status == "unrecognized"


def test_excel_marks_invalid_source_as_not_found(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "주소 미상"
    wb.save(input_path)

    stats = process_workbook(input_path, output_path, source_col="H", target_col="I", status_col="N")

    result_wb = openpyxl.load_workbook(output_path)
    result_ws = result_wb.active
    assert result_ws["I2"].value is None
    assert result_ws["N2"].value == STATUS_NOT_FOUND
    assert stats["missing"] == 1


def test_mark_missing_without_api_keys_raises(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호"
    wb.save(input_path)

    try:
        process_workbook(
            input_path,
            output_path,
            source_col="H",
            target_col="I",
            status_col="N",
            mark_missing=True,
        )
    except RuntimeError as exc:
        assert "API key" in str(exc)
    else:
        raise AssertionError("expected RuntimeError")


def test_provider_none_with_mark_missing_does_not_mark_searchable_rows(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제1층 제101호"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="N",
        provider="none",
        mark_missing=True,
    )

    result_wb = openpyxl.load_workbook(output_path)
    result_ws = result_wb.active
    assert result_ws["I2"].value == "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    assert result_ws["N2"].value is None
    assert stats["missing"] == 0


class _FakeJuso:
    def __init__(self, total_count: int, error: bool = False):
        self.total_count = total_count
        self.error = error

    def search(self, query: str, count: int = 5):
        first = {"errorCode": "E"} if self.error else {}
        return SearchResult("juso", self.total_count, first, {})


class _FakeEpost:
    def __init__(self, total_count: int, error: bool = False):
        self.total_count = total_count
        self.error = error

    def search(self, query: str, search_se: str = "road", count: int = 5):
        first = {"returnCode": "30"} if self.error else {}
        return SearchResult("epost", self.total_count, first, "")


class _FallbackEpost:
    def __init__(self):
        self.queries = []

    def search(self, query: str, search_se: str = "road", count: int = 5):
        self.queries.append((query, search_se))
        if query == "하우3길 22":
            return SearchResult("epost", 1, {}, "")
        return SearchResult("epost", 0, {"returnCode": "01"}, "")


def test_verify_marks_ambiguous_when_multiple_results():
    assert _verify("서울특별시 강남구 테헤란로 152", "road", _FakeJuso(2), None) == "ambiguous"


def test_verify_marks_ambiguous_before_single_match():
    assert _verify("서울특별시 강남구 테헤란로 152", "road", _FakeJuso(1), _FakeEpost(2)) == "ambiguous"


def test_verify_ignores_one_provider_error_when_another_provider_works():
    assert _verify("서울특별시 강남구 테헤란로 152", "road", _FakeJuso(1), _FakeEpost(0, error=True)) == "verified"


def test_verify_raises_when_all_configured_providers_error():
    try:
        _verify("서울특별시 강남구 테헤란로 152", "road", _FakeJuso(0, error=True), None)
    except RuntimeError as exc:
        assert "API errors" in str(exc)
    else:
        raise AssertionError("expected RuntimeError")


class _RaisingJuso:
    key = "test-key"

    def search(self, query: str, count: int = 5):
        raise requests.ConnectionError("network down")


def test_verify_tolerates_juso_transport_error_when_epost_works():
    assert _verify("서울특별시 강남구 테헤란로 152", "road", _RaisingJuso(), _FakeEpost(1)) == "verified"


def test_verify_raises_when_transport_error_hits_the_only_provider():
    try:
        _verify("서울특별시 강남구 테헤란로 152", "road", _RaisingJuso(), None)
    except RuntimeError as exc:
        assert "API errors" in str(exc)
    else:
        raise AssertionError("expected RuntimeError")


def test_process_workbook_reuses_verification_for_repeated_addresses(tmp_path, monkeypatch):
    calls: list[str] = []

    class _CountingJuso:
        key = "test-key"

        def search(self, query: str, count: int = 5):
            calls.append(query)
            return SearchResult("juso", 1, {}, {})

    monkeypatch.setattr("address_cleaner.excel.JusoClient", _CountingJuso)
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    ws["H3"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="N",
        provider="juso",
        mark_missing=True,
    )

    assert stats["verified"] == 2
    assert len(calls) == 1


def test_process_workbook_marks_m_column_when_juso_returns_no_results(tmp_path, monkeypatch):
    class _MissingJuso:
        key = "test-key"

        def search(self, query: str, count: int = 5):
            return SearchResult("juso", 0, {}, {})

    monkeypatch.setattr("address_cleaner.excel.JusoClient", _MissingJuso)
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        provider="juso",
        mark_missing=True,
    )

    result_ws = openpyxl.load_workbook(output_path).active
    assert result_ws["M1"].value == "주소검색결과"
    assert result_ws["M2"].value == STATUS_NOT_FOUND
    assert stats["missing"] == 1


def test_process_workbook_marks_m_column_when_juso_returns_multiple_results(tmp_path, monkeypatch):
    class _AmbiguousJuso:
        key = "test-key"

        def search(self, query: str, count: int = 5):
            return SearchResult("juso", 2, {}, {})

    monkeypatch.setattr("address_cleaner.excel.JusoClient", _AmbiguousJuso)
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        provider="juso",
        mark_missing=True,
    )

    result_ws = openpyxl.load_workbook(output_path).active
    assert result_ws["M1"].value == "주소검색결과"
    assert result_ws["M2"].value == STATUS_AMBIGUOUS
    assert stats["ambiguous"] == 1


def test_verify_falls_back_to_compact_epost_query():
    epost = _FallbackEpost()
    result = _verify("경기도 파주시 하우3길 22 정우펠리스 제303동 제101호", "road", None, epost)
    assert result == "verified"
    assert epost.queries == [
        ("경기도 파주시 하우3길 22 정우펠리스 제303동 제101호", "road"),
        ("하우3길 22", "road"),
    ]


def test_status_labels_are_user_facing_korean():
    assert STATUS_NOT_FOUND == "검색주소없음"
    assert STATUS_AMBIGUOUS == "2건이상검색"


def test_cli_prints_clean_error_for_missing_api_keys(tmp_path, capsys):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17"
    wb.save(input_path)

    exit_code = main([
        "excel",
        str(input_path),
        "-o",
        str(output_path),
        "--source-col",
        "H",
        "--target-col",
        "I",
        "--status-col",
        "N",
        "--mark-missing",
    ])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")


def test_registry_subcommand_delegates_to_registry_refiner(tmp_path, capsys, monkeypatch):
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "대상 임대차계약 주소"
    ws["A2"] = "서울특별시 강남구 역삼동 123-4 101호"
    wb.save(input_path)
    monkeypatch.delenv("JUSO_CONFIRM_KEY", raising=False)
    monkeypatch.delenv("JUSO_CONFM_KEY", raising=False)
    monkeypatch.delenv("JUSO_API_KEY", raising=False)
    monkeypatch.delenv("CONFM_KEY", raising=False)

    exit_code = main(["registry", str(input_path), "-o", str(tmp_path / "out")])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "JUSO_CONFM_KEY" in captured.err


def test_registry_address_refine_console_script_is_kept_for_compatibility():
    pyproject = Path("pyproject.toml").read_text(encoding="utf-8")

    assert 'registry-address-refine = "address_cleaner.registry.cli:main"' in pyproject


def test_preprocess_cuts_repeated_legacy_sido_names():
    assert (
        preprocess_raw_address("강원도 춘천시 중앙로 1 강원도 춘천시 중앙로 1")
        == "강원도 춘천시 중앙로 1"
    )


def test_road_address_without_district_is_searchable():
    result = normalize_for_search("테헤란로 152 강남파이낸스센터 10층")
    assert result.kind == "road"
    assert result.query == "테헤란로 152 강남파이낸스센터"
    assert result.searchable


def test_empty_source_rows_keep_status_blank(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["H1"] = "원주소"
    ws["H2"] = "경기도 파주시 야당동 57-17 정우펠리스 제303동 제101호"
    ws["A3"] = "서식만 남은 행"
    wb.save(input_path)

    stats = process_workbook(
        input_path,
        output_path,
        source_col="H",
        target_col="I",
        status_col="M",
        provider="none",
    )

    result_ws = openpyxl.load_workbook(output_path).active
    assert result_ws["M2"].value in (None, "")
    assert result_ws["M3"].value in (None, "")
    assert stats["empty"] == 1


def test_epost_search_retries_transient_errors(monkeypatch):
    from address_cleaner import clients

    calls = {"n": 0}

    class _Resp:
        text = "<root><returnCode>00</returnCode><totalCount>1</totalCount></root>"

        def raise_for_status(self):
            return None

    def fake_get(url, params=None, timeout=None):
        calls["n"] += 1
        if calls["n"] == 1:
            raise requests.ConnectionError("일시적 네트워크 오류")
        return _Resp()

    monkeypatch.setattr(clients.requests, "get", fake_get)
    monkeypatch.setattr(clients.time, "sleep", lambda s: None)

    result = clients.KoreaPostRoadNameClient(key="k").search("테헤란로 152")

    assert result.total_count == 1
    assert calls["n"] == 2


def test_duplex_floor_detail_with_parenthetical_is_removed():
    # 실제 전산 입력 실패 사례: '제1(상층하층)층'이 검색어에 남아 우편번호 확정 실패.
    result = normalize_for_search("경기도 고양시 일산동구 중산동 78-7 시크릿타운 제비동 제1(상층하층)층 제101호")
    assert result.query == "경기도 고양시 일산동구 중산동 78-7 시크릿타운 제비동 제101호"
    assert result.kind == "lot"
