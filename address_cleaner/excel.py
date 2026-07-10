from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
import json
from pathlib import Path
from typing import Any, Literal

import openpyxl
import requests

from .clients import JusoClient, KoreaPostRoadNameClient, SearchResult
from .history import VerifyHistory
from .juso_search import RateLimiter
from .normalizer import base_for_search, compact_for_epost, normalize_for_search
from .regions import AdminDict

# 등기 모드의 검색어 빌딩블록을 일반 모드 검증에도 재사용한다.
from .registry.normalize import district_key, lot_variants, parse_lot_addr


ProviderMode = Literal["none", "juso", "epost", "both"]
STATUS_NOT_FOUND = "검색주소없음"
STATUS_AMBIGUOUS = "2건이상검색"

# 로컬 정제 단계에서 검색 불가 판정된 사유의 사람용 설명 (검증상세 컬럼용)
LOCAL_STATUS_KO = {
    "invalid_marker": "원주소가 '미상' 등 무효 표기",
    "malformed": "주소 형식이 아님 (행정구역/번지 없음)",
    "unrecognized": "지번/도로명 골격을 찾지 못함",
}

# 등기 모드와 동일 기준의 병렬 검증 설정. 워커가 늘어도 provider별 전역
# 레이트리미터가 초당 호출 수를 MAX_REQ_PER_SEC 이하로 묶는다.
DEFAULT_WORKERS = 8
MAX_REQ_PER_SEC = 10.0

# (판정, 사람이 읽을 검증 상세, 교정 후보)
VerifyOutcome = tuple[str, str, "dict[str, str] | None"]

Verdict = Literal["verified", "ambiguous", "missing"]


@dataclass(frozen=True)
class VerifyResult:
    """verify_address()의 결과.

    - verdict: verified(1건 확정) / ambiguous(2건 이상) / missing(0건 또는 검색어 없음)
    - detail: 사람이 읽을 검증 상세 (검색 경로/건수/표준주소/우편번호)
    - correction: 원문 그대로는 안 되고 골격/지번 변형으로만 통한 경우의 교정 후보
    """

    verdict: Verdict
    detail: str
    correction: dict[str, str] | None = None

    @property
    def verified(self) -> bool:
        return self.verdict == "verified"


def col_to_index(col: str) -> int:
    col = col.strip().upper()
    if not col:
        # 빈 문자열이 0(존재하지 않는 열)으로 조용히 통과하면 저장 시점에야 깨진다.
        raise ValueError("Invalid Excel column: (empty)")
    value = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid Excel column: {col}")
        value = value * 26 + ord(ch) - ord("A") + 1
    return value


def process_workbook(
    input_path: str | Path,
    output_path: str | Path,
    source_col: str = "H",
    target_col: str = "I",
    status_col: str | None = None,
    detail_col: str | None = None,
    provider: ProviderMode = "both",
    mark_missing: bool = False,
    header: bool = True,
    admin_dict: AdminDict | None = None,
    history: VerifyHistory | None = None,
    corrections_path: str | Path | None = None,
    workers: int = DEFAULT_WORKERS,
) -> dict[str, int]:
    """엑셀을 정제·검증한다. openpyxl 워크시트는 스레드 안전하지 않으므로
    셀 접근과 SQLite 이력 접근은 메인 스레드로 한정하는 2패스 구조다.

    1패스: 행 순회 — 정제·로컬 판정(빈 행/무효/법정동 사전)은 즉시 확정하고,
           API 검증이 필요한 고유 (검색어, 종류)만 모은다 (이력 재사용분 제외).
    병렬:  고유 검색어만 ThreadPoolExecutor(workers)로 _verify. workers=1이면 직렬.
    2패스: 행 순회하며 결과를 셀에 기록하고 stats/교정 후보를 집계한다.
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    source_idx = col_to_index(source_col)
    target_idx = col_to_index(target_col)
    status_idx = col_to_index(status_col) if status_col else None
    detail_idx = col_to_index(detail_col) if detail_col else None
    if detail_idx and not status_idx:
        raise RuntimeError("--detail-col requires --status-col")

    if header:
        ws.cell(row=1, column=target_idx).value = "주소검색어"
        if status_idx:
            ws.cell(row=1, column=status_idx).value = "주소검색결과"
        if detail_idx:
            ws.cell(row=1, column=detail_idx).value = "주소검증상세"

    juso = JusoClient() if provider in ("juso", "both") else None
    epost = KoreaPostRoadNameClient() if provider in ("epost", "both") else None
    if juso is not None and not juso.key:
        juso = None
    if epost is not None and not epost.key:
        epost = None
    if mark_missing and provider != "none" and juso is None and epost is None:
        raise RuntimeError(
            "At least one API key is required when --mark-missing validates provider results"
        )
    # 커넥션 재사용을 위한 공유 Session (등기 모드와 동일 패턴 — Session은 스레드 안전).
    session = requests.Session()
    if juso is not None:
        juso.session = session
    if epost is not None:
        epost.session = session

    stats = {
        "total": 0,
        "road": 0,
        "lot": 0,
        "empty": 0,
        "invalid": 0,
        "missing": 0,
        "ambiguous": 0,
        "verified": 0,
        "history_reused": 0,
        "verdict_changed": 0,
        "correction_candidates": 0,
    }
    start_row = 2 if header else 1
    # 같은 원주소가 여러 행에 반복되는 파일이 흔해서 검증 결과를 재사용한다.
    verify_cache: dict[tuple[str, str], VerifyOutcome] = {}
    # 교정 후보: (원문, 통한 검색어) 단위로 모아 어떤 행들에서 나왔는지 누적한다.
    corrections: dict[tuple[str, str], dict[str, Any]] = {}

    # 1패스 (메인 스레드): 정제·로컬 판정 확정 + 검증 대상 수집.
    local_rows: list[tuple[int, str, str]] = []  # (행, 상태, 검증상세) — 확정분
    verify_rows: list[tuple[int, tuple[str, str]]] = []  # (행, 검증 키) — 대기분
    to_verify: list[tuple[str, str]] = []  # API 호출이 필요한 고유 키 (행 순서 유지)
    queued: set[tuple[str, str]] = set()  # to_verify 중복 검사용
    for row in range(start_row, ws.max_row + 1):
        raw = ws.cell(row=row, column=source_idx).value
        normalized = normalize_for_search(raw)
        ws.cell(row=row, column=target_idx).value = normalized.query
        stats["total"] += 1
        stats[normalized.kind if normalized.kind in stats else "invalid"] += 1

        if not status_idx:
            continue
        if normalized.kind == "empty":
            # 원주소가 비어 있는 행(서식만 남은 말미 행 포함)은 검토 대상이
            # 아니므로 상태를 비워 둬 후단 자동화가 불량 주소와 혼동하지 않게 한다.
            local_rows.append((row, "", ""))
        elif not normalized.searchable:
            local_rows.append(
                (
                    row,
                    STATUS_NOT_FOUND,
                    LOCAL_STATUS_KO.get(normalized.status, normalized.status),
                )
            )
            stats["missing"] += 1
        elif admin_dict is not None and (
            dict_reason := _admin_combo_missing(
                normalized.query, normalized.kind, admin_dict
            )
        ):
            # 법정동 사전 오프라인 검증: API 호출 전에 실존하지 않는 행정구역을 거른다.
            local_rows.append((row, STATUS_NOT_FOUND, dict_reason))
            stats["missing"] += 1
        elif mark_missing and (juso is not None or epost is not None):
            cache_key = (normalized.query, normalized.kind)
            verify_rows.append((row, cache_key))
            if cache_key in verify_cache or cache_key in queued:
                continue
            # 이력 조회는 메인 스레드에서 미리 끝내 병렬 대상에서 제외한다.
            fresh = history.fresh(*cache_key) if history else None
            if fresh is not None:
                verify_cache[cache_key] = (
                    fresh.verdict,
                    f"{fresh.detail} (이력 재사용 {fresh.checked_at[:10]})",
                    None,
                )
                stats["history_reused"] += 1
            else:
                to_verify.append(cache_key)
                queued.add(cache_key)
        else:
            local_rows.append((row, "", ""))

    # 병렬 구간: 고유 (검색어, 종류)만 검증한다. 셀/이력은 건드리지 않는 순수 계산.
    if to_verify:
        juso_limiter = RateLimiter(MAX_REQ_PER_SEC)
        epost_limiter = RateLimiter(MAX_REQ_PER_SEC)

        def verify_one(key: tuple[str, str]) -> VerifyOutcome:
            return _verify(
                key[0],
                key[1],
                juso,
                epost,
                juso_limiter=juso_limiter,
                epost_limiter=epost_limiter,
            )

        outcomes: dict[tuple[str, str], VerifyOutcome] = {}
        if workers > 1 and len(to_verify) > 1:
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {executor.submit(verify_one, key): key for key in to_verify}
                try:
                    for future in as_completed(futures):
                        outcomes[futures[future]] = future.result()
                except BaseException:
                    # 전 provider 오류(RuntimeError) 등이 나면 남은 작업을 취소하고
                    # 그대로 전파해 "키가 죽었는데 전량 검색주소없음" 사고를 막는다.
                    for pending in futures:
                        pending.cancel()
                    raise
        else:
            outcomes = {key: verify_one(key) for key in to_verify}

        # history 기록·판정 변경 감지는 메인 스레드에서 수행한다
        # (sqlite3 기본 연결은 생성 스레드 전용이므로 이 구조가 락 없이 안전).
        for cache_key in to_verify:
            verification, verify_detail, correction = outcomes[cache_key]
            if history is not None:
                previous = history.latest(*cache_key)
                history.record(cache_key[0], cache_key[1], verification, verify_detail)
                if previous is not None and previous.verdict != verification:
                    # 행정구역 개편·건물 멸실 등의 신호: 사람이 봐야 한다.
                    verify_detail += f"; ⚠ 판정 변경: {previous.checked_at[:10]} {previous.verdict} → {verification}"
                    stats["verdict_changed"] += 1
            verify_cache[cache_key] = (verification, verify_detail, correction)

    # 2패스 (메인 스레드): 결과를 셀에 기록하고 stats/교정 후보를 집계한다.
    if status_idx:
        for row, status, verify_detail in local_rows:
            ws.cell(row=row, column=status_idx).value = status
            if detail_idx:
                ws.cell(row=row, column=detail_idx).value = verify_detail
        for row, cache_key in verify_rows:
            verification, verify_detail, correction = verify_cache[cache_key]
            if correction is not None:
                key = (correction["original"], correction["working"])
                entry = corrections.setdefault(key, {**correction, "rows": []})
                entry["rows"].append(row)
            status = ""
            if verification == "verified":
                stats["verified"] += 1
            elif verification == "ambiguous":
                status = STATUS_AMBIGUOUS
                stats["ambiguous"] += 1
            else:
                status = STATUS_NOT_FOUND
                stats["missing"] += 1
            ws.cell(row=row, column=status_idx).value = status
            if detail_idx:
                ws.cell(row=row, column=detail_idx).value = verify_detail

    wb.save(output_path)

    stats["correction_candidates"] = len(corrections)
    if corrections_path is not None:
        # 사람이 검토해 typo-rules나 원주소 수정에 반영할 교정 후보 리포트.
        report = {
            "input": str(input_path),
            "generated": datetime.now().isoformat(timespec="seconds"),
            "candidates": list(corrections.values()),
        }
        Path(corrections_path).write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    return stats


def collect_feedback(
    input_path: str | Path,
    source_col: str = "H",
    target_col: str = "I",
    result_col: str = "M",
    ps_detail_col: str | None = "N",
    header: bool = True,
) -> dict[str, Any]:
    """파워쉘 처리결과 엑셀에서 실패 행을 모아 재정제 리포트를 만든다.

    M열이 '실패(...)'인 행에 대해 현재 규칙으로 원주소를 다시 정제해 보고,
    기존 I열과 달라졌으면(=규칙이 그동안 개선됐으면) I열 갱신 후보로 표시한다.
    여전히 같은 검색어가 나오는 행은 새 정제 규칙이 필요한 사례다.
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    source_idx = col_to_index(source_col)
    target_idx = col_to_index(target_col)
    result_idx = col_to_index(result_col)
    ps_detail_idx = col_to_index(ps_detail_col) if ps_detail_col else None

    rows: list[dict[str, Any]] = []
    by_result: dict[str, int] = {}
    requery_changed = 0
    for row in range(2 if header else 1, ws.max_row + 1):
        result = str(ws.cell(row=row, column=result_idx).value or "")
        if not result.startswith("실패"):
            continue
        source = ws.cell(row=row, column=source_idx).value
        current_query = str(ws.cell(row=row, column=target_idx).value or "")
        normalized = normalize_for_search(source)
        changed = bool(normalized.query) and normalized.query != current_query
        if changed:
            requery_changed += 1
        by_result[result] = by_result.get(result, 0) + 1
        rows.append(
            {
                "row": row,
                "result": result,
                "source": str(source or ""),
                "currentQuery": current_query,
                "psDetail": str(ws.cell(row=row, column=ps_detail_idx).value or "")
                if ps_detail_idx
                else "",
                "requery": normalized.query,
                "requeryChanged": changed,
                "localStatus": normalized.status,
            }
        )
    return {
        "input": str(input_path),
        "generated": datetime.now().isoformat(timespec="seconds"),
        "failures": len(rows),
        "byResult": by_result,
        "requeryChanged": requery_changed,
        "rows": rows,
    }


def _admin_combo_missing(query: str, kind: str, admin_dict: AdminDict) -> str:
    """주소의 행정구역 조합이 법정동 사전에 없으면 사유 문자열, 있으면 빈 문자열.

    행정동 표기(예: 신정3동)는 법정동 사전에 없어 거짓 양성이 날 수 있으므로
    이 검사는 표시까지만 하고, 최종 판단은 사람/API 검증에 맡긴다.
    """
    if kind == "lot":
        lot = parse_lot_addr(query)
        combo = " ".join(
            p
            for p in [
                lot["sido"],
                lot["city"],
                lot["sigungu"],
                lot.get("eupmyeon", ""),
                lot["dong"],
            ]
            if p
        )
    else:
        combo = district_key(query)
    if combo and not admin_dict.contains(combo):
        return f"법정동 사전에 없는 행정구역: {combo}"
    return ""


def _result_note(provider_label: str, result: SearchResult) -> str:
    if result.total_count == 1:
        road = result.first.get("roadAddr") or result.first.get("lnmAdres") or ""
        zip_no = result.first.get("zipNo") or result.first.get("zipNo1") or ""
        tail = (
            f": {road}" + (f" (우){zip_no}" if zip_no else "") if road or zip_no else ""
        )
        return f"{provider_label} 1건{tail}"
    return f"{provider_label} {result.total_count}건"


def verify_address(
    query: str,
    kind: str,
    *,
    juso: JusoClient | None = None,
    epost: KoreaPostRoadNameClient | None = None,
) -> VerifyResult:
    """주소 검색어 1건을 juso.go.kr/우체국 API로 검증한다 (엑셀 없이 사용 가능).

    query는 normalize_for_search()가 만든 검색어(.query), kind는 그 종류(.kind:
    "road"|"lot")를 그대로 넘기면 된다. 클라이언트를 넘기지 않으면 환경변수 키로
    기본 클라이언트를 만들며, 사용 가능한 provider가 하나도 없으면 RuntimeError.

    >>> normalized = normalize_for_search("경기도 파주시 야당동 57-17 ...")
    >>> result = verify_address(normalized.query, normalized.kind)
    >>> result.verdict
    'verified'
    """
    if juso is None and epost is None:
        juso = JusoClient()
        epost = KoreaPostRoadNameClient()
        if not juso.key:
            juso = None
        if not epost.key:
            epost = None
        if juso is None and epost is None:
            raise RuntimeError(
                "At least one API key is required "
                "(JUSO_CONFIRM_KEY or EPOST_SERVICE_KEY)"
            )
    verdict, detail, correction = _verify(query, kind, juso, epost)
    return VerifyResult(verdict, detail, correction)


def _verify(
    query: str,
    kind: str,
    juso: JusoClient | None,
    epost: KoreaPostRoadNameClient | None,
    *,
    juso_limiter: RateLimiter | None = None,
    epost_limiter: RateLimiter | None = None,
) -> tuple[Verdict, str, dict[str, str] | None]:
    """(판정, 사람이 읽을 검증 상세, 교정 후보) 반환.

    입력 → 출력만 있는 순수 계산이라 그대로 병렬화 단위가 된다. 호출 페이싱은
    provider별 공용 RateLimiter가 맡는다 (직렬·병렬 모두 같은 초당 상한).

    Juso는 상세 포함 검색이 0건이면 골격(시도~지번/건물번호)으로 한 번 더 검색한다.
    상세 표기('제비동 제101호' 등) 때문에 멀쩡한 주소가 불량 처리되는 것을 막고,
    골격조차 0건인 진짜 불량과 구분되도록 상세에 검색 경로를 남긴다.

    교정 후보: 원문 그대로는 안 되지만 골격/지번 변형으로는 통한 경우,
    '원문 ↔ 통한 검색어' 쌍을 반환해 교정 규칙 후보 리포트에 누적할 수 있게 한다.
    """
    if not query:
        return "missing", "", None
    results: list[SearchResult] = []
    notes: list[str] = []
    correction: dict[str, str] | None = None
    if juso is not None:
        juso_queries = [("상세포함", query)]
        base = base_for_search(query, kind)
        if base and base != query:
            juso_queries.append(("골격", base))
        for label, juso_query in juso_queries:
            if juso_limiter is not None:
                juso_limiter.wait()
            try:
                result = juso.search(juso_query, count=5)
            except Exception as exc:
                # 재시도 후에도 남은 전송 오류는 다른 provider 결과로 판정을 이어가고,
                # 모든 provider가 오류면 아래에서 실행을 중단한다.
                result = SearchResult(
                    "juso",
                    0,
                    {"errorCode": "transport_error", "errorMessage": str(exc)},
                    "",
                )
            if result.has_error:
                notes.append(f"JUSO[{label}] 오류")
                break
            notes.append(_result_note(f"JUSO[{label}]", result))
            if result.total_count != 0:
                if label == "골격" and result.total_count == 1:
                    # 상세부 표기가 검색을 깨뜨린 사례 → 교정 후보로 수확
                    correction = {
                        "type": "상세부제거",
                        "original": query,
                        "working": base,
                        "resolved": result.first.get("roadAddr") or "",
                        "zip": result.first.get("zipNo") or "",
                    }
                break
        results.append(result)
        # 골격까지 0건인 지번주소는 붙여 쓴 지번(5717→57-17) 변형으로 후보를 찾아
        # 사람이 보완할 수 있게 제안만 남긴다. 주소를 자동으로 바꾸지는 않는다.
        if not result.has_error and result.total_count == 0 and kind == "lot":
            for variant in lot_variants(base or query)[:3]:
                if juso_limiter is not None:
                    juso_limiter.wait()
                try:
                    variant_result = juso.search(variant, count=5)
                except Exception:
                    break
                if not variant_result.has_error and variant_result.total_count == 1:
                    road = variant_result.first.get("roadAddr") or ""
                    zip_no = variant_result.first.get("zipNo") or ""
                    notes.append(
                        f"지번 변형 후보 1건: {variant}"
                        + (f" → {road}" if road else "")
                        + (f" (우){zip_no}" if zip_no else "")
                    )
                    correction = {
                        "type": "지번변형",
                        "original": base or query,
                        "working": variant,
                        "resolved": road,
                        "zip": zip_no,
                    }
                    break
    if epost is not None:
        search_se = "road" if kind == "road" else "dong"
        epost_queries = [query]
        compact_query = compact_for_epost(query, kind)
        if compact_query and compact_query not in epost_queries:
            epost_queries.append(compact_query)
        for epost_query in epost_queries:
            if epost_limiter is not None:
                epost_limiter.wait()
            try:
                result = epost.search(epost_query, search_se=search_se, count=5)
            except Exception as exc:
                result = SearchResult(
                    "epost",
                    0,
                    {"returnCode": "transport_error", "returnMessage": str(exc)},
                    "",
                )
            results.append(result)
            if not result.has_error and result.total_count != 0:
                break
        notes.append(
            "EPOST 오류" if result.has_error else _result_note("EPOST", result)
        )
    usable_results = [result for result in results if not result.has_error]
    if not usable_results and results:
        raise RuntimeError(
            "Address validation providers returned API errors; check API keys before marking missing addresses"
        )
    detail = "; ".join(notes)
    if any(result.total_count >= 2 for result in usable_results):
        return "ambiguous", detail, correction
    if any(result.total_count == 1 for result in usable_results):
        return "verified", detail, correction
    return "missing", detail, correction
