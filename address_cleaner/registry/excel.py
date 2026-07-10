"""엑셀 입출력과 정제 파이프라인(refine)."""

from __future__ import annotations

import json
import re
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import openpyxl
import requests

from .juso import (
    RateLimiter,
    cache_lock,
    candidate_id,
    first_pass_status,
    juso_query,
    load_cache,
    make_queries,
    save_cache,
    score_candidate,
    set_rate_limiter,
)
from .normalize import (
    building_name,
    has_extra_parcels,
    norm,
    original_is_under_specified,
    parse_lot_addr,
    strip_detail,
    suffix_dong,
    suffix_ho,
    unit_extract,
    unit_out_of_range,
)

# 콜드 캐시 첫 실행의 벽시계 시간을 줄이는 병렬 처리 설정.
# 워커가 늘어도 전역 레이트리미터가 초당 호출을 MAX_REQ_PER_SEC 이하로 묶는다.
DEFAULT_WORKERS = 8
MAX_REQ_PER_SEC = 10.0

JUSO_COLUMNS = [
    "JUSO_판정",
    "JUSO_검색어",
    "JUSO_총건수",
    "JUSO_도로명주소_1",
    "JUSO_지번주소_1",
    "JUSO_우편번호_1",
    "JUSO_시군구코드_1",
    "JUSO_결과상위5",
]
JUSO2_COLUMNS = [
    "JUSO_2차판정",
    "JUSO_2차점수",
    "JUSO_2차점수차",
    "JUSO_추천도로명주소",
    "JUSO_추천지번주소",
    "JUSO_추천우편번호",
    "JUSO_추천시군구코드",
    "JUSO_추천근거",
    "JUSO_2차검색어",
    "JUSO_2차후보상위5",
]
REGISTRY_COLUMNS = [
    "등기_부동산구분",
    "등기_시도",
    "등기_시군구",
    "등기_법정동",
    "등기_지번",
    "등기_지번주소",
    "등기_건물명_원문추정",
    "등기_동",
    "등기_층",
    "등기_호",
    "등기_검색문구",
    "등기_조회전략",
    "등기_검토등급",
    "등기_검토사유",
]
WHOLE_COLUMNS = [
    "등기소_전체검색어_동호포함",
    "등기소_전체검색_보조검색어",
    "등기소_전체검색_상태",
]
FINAL_COLUMNS = ["최종 검색용 주소", "주소검토결과"]


def ensure_columns(ws: Any, names: list[str]) -> dict[str, int]:
    headers = {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    for name in names:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(1, col).value = name
            headers[name] = col
    return headers


def get_cell(ws: Any, row: int, headers: dict[str, int], name: str) -> Any:
    col = headers.get(name)
    return ws.cell(row, col).value if col else ""


def set_row_values(
    ws: Any, row: int, headers: dict[str, int], names: list[str], values: list[Any]
) -> None:
    for name, value in zip(names, values):
        ws.cell(row, headers[name]).value = value


def add_or_replace_sheet(wb: Any, name: str) -> Any:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def refine(
    input_file: Path,
    output_dir: Path,
    key: str,
    cache_file: Path | None = None,
    verbose: bool = True,
    workers: int = DEFAULT_WORKERS,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_file or (output_dir / f"{input_file.stem}_juso_cache.json")
    cache = load_cache(cache_file)
    workers = max(1, int(workers))

    wb = openpyxl.load_workbook(input_file)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    headers = ensure_columns(
        ws,
        JUSO_COLUMNS + JUSO2_COLUMNS + REGISTRY_COLUMNS + WHOLE_COLUMNS + FINAL_COLUMNS,
    )

    required = ["대상 임대차계약 주소"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError("필수 컬럼 없음: " + ", ".join(missing))

    def source_addr(row: int) -> str:
        # Prefer a human-reviewed final address when present; otherwise use the raw lease address.
        # This lets the same refiner work on both already-refined workbooks and raw 통합 sheets.
        return norm(
            get_cell(ws, row, headers, "최종주소")
            or get_cell(ws, row, headers, "대상 임대차계약 주소")
        )

    session = requests.Session()
    limiter = RateLimiter(MAX_REQ_PER_SEC) if workers > 1 else None

    # 1차 Juso 검사: source address unique 기준 캐시.
    final_addresses = [source_addr(r) for r in range(2, ws.max_row + 1)]
    unique_addresses = sorted({x for x in final_addresses if x})
    first_items: dict[str, Any] = {}

    def first_pass(addr: str) -> dict[str, Any]:
        item_key = f"first:{addr}"
        with cache_lock():
            cached = cache.get(item_key)
        if cached is not None:
            return cached
        full = juso_query(session, key, addr, cache, count=5, preserve_commas=True)
        stripped = strip_detail(addr)
        normalized = (
            juso_query(session, key, stripped, cache, count=5, preserve_commas=True)
            if full["total"] == 0 and stripped and stripped != addr
            else None
        )
        status, best = first_pass_status(full, normalized)
        item = {
            "address": addr,
            "normalizedKeyword": stripped,
            "full": full,
            "normalized": normalized,
            "status": status,
            "best": best,
        }
        with cache_lock():
            cache[item_key] = item
        return item

    set_rate_limiter(limiter)
    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(first_pass, addr): addr for addr in unique_addresses
            }
            for done, future in enumerate(as_completed(futures), 1):
                first_items[futures[future]] = future.result()
                if done % 50 == 0:
                    save_cache(cache_file, cache)
                    if verbose:
                        print(f"Juso 1차 {done}/{len(unique_addresses)}", flush=True)
    finally:
        # 중간에 실패해도 지금까지 받은 API 응답은 보존한다.
        set_rate_limiter(None)
        save_cache(cache_file, cache)

    first_counter: Counter[str] = Counter()
    target_rows: list[int] = []
    for r in range(2, ws.max_row + 1):
        addr = source_addr(r)
        item = first_items.get(addr)
        values: list[Any]
        if not item:
            values = ["최종주소_공백", "", 0, "", "", "", "", ""]
        else:
            best = item["best"]
            rows = best.get("rows") or []
            first = rows[0] if rows else {}
            values = [
                item["status"],
                best.get("keyword", ""),
                best.get("total", 0),
                first.get("roadAddr", ""),
                first.get("jibunAddr", ""),
                first.get("zipNo", ""),
                first.get("admCd", ""),
                "\n".join(
                    f"{i + 1}. {x.get('roadAddr', '')} / {x.get('jibunAddr', '')}"
                    for i, x in enumerate(rows[:5])
                ),
            ]
            if item["status"] == "검색불가" or item["status"].startswith("다중검출"):
                target_rows.append(r)
        first_counter[values[0]] += 1
        set_row_values(ws, r, headers, JUSO_COLUMNS, values)

    # 2차 좁힘.
    second_counter: Counter[str] = Counter()
    # openpyxl은 스레드 안전이 아니므로 셀 읽기는 메인스레드에서 먼저 모으고,
    # API 검색·스코어링만 워커에서 병렬로 돌린 뒤 결과를 메인스레드가 직렬로 쓴다.
    second_inputs = [
        (
            r,
            get_cell(ws, r, headers, "대상 임대차계약 주소"),
            source_addr(r),
            get_cell(ws, r, headers, "도로명_완성"),
            get_cell(ws, r, headers, "지번_완성"),
        )
        for r in target_rows
    ]

    def second_pass(raw: Any, final: str, roadc: Any, jibunc: Any) -> list[Any]:
        queries = make_queries(raw, final, roadc, jibunc)
        candidates: dict[str, dict[str, Any]] = {}
        for q in queries:
            res = juso_query(session, key, q, cache, count=30)
            for row in res.get("rows") or []:
                cid = candidate_id(row)
                candidates.setdefault(cid, {"row": row, "queries": []})[
                    "queries"
                ].append({"keyword": res["keyword"], "total": res["total"]})
        scored: list[tuple[int, str, dict[str, Any], list[str]]] = []
        for cid, obj in candidates.items():
            score, reasons = score_candidate(
                obj["row"], raw, final, roadc, jibunc, obj["queries"]
            )
            scored.append((score, cid, obj, reasons))
        scored.sort(key=lambda x: x[0], reverse=True)
        if not scored:
            return [
                "수동검토",
                0,
                0,
                "",
                "",
                "",
                "",
                "후보없음",
                "\n".join(queries),
                "",
            ]
        score, _cid, obj, reasons = scored[0]
        second = scored[1][0] if len(scored) > 1 else 0
        margin = score - second
        row = obj["row"]
        if score >= 80 and margin >= 15:
            decision = "자동추천_높음"
        elif score >= 65 and margin >= 10:
            decision = "자동추천_중간"
        elif score >= 50:
            decision = "후보1_검토"
        else:
            decision = "수동검토"
        return [
            decision,
            score,
            margin,
            row.get("roadAddr", ""),
            row.get("jibunAddr", ""),
            row.get("zipNo", ""),
            row.get("admCd", ""),
            "; ".join(reasons),
            "\n".join(f"{x['keyword']} ({x['total']})" for x in obj["queries"][:6]),
            "\n".join(
                f"{i + 1}. [{s}] {o['row'].get('roadAddr', '')} / {o['row'].get('jibunAddr', '')}"
                for i, (s, _, o, _) in enumerate(scored[:5])
            ),
        ]

    second_values: dict[int, list[Any]] = {}
    set_rate_limiter(limiter)
    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            second_futures = {
                executor.submit(second_pass, raw, final, roadc, jibunc): r
                for (r, raw, final, roadc, jibunc) in second_inputs
            }
            for done, second_future in enumerate(as_completed(second_futures), 1):
                second_values[second_futures[second_future]] = second_future.result()
                if done % 50 == 0:
                    save_cache(cache_file, cache)
                    if verbose:
                        print(f"Juso 2차 {done}/{len(target_rows)}", flush=True)
    finally:
        set_rate_limiter(None)
        save_cache(cache_file, cache)

    for r in target_rows:
        values = second_values[r]
        second_counter[values[0]] += 1
        set_row_values(ws, r, headers, JUSO2_COLUMNS, values)

    # 등기소 전체검색어 생성.
    final_counter: Counter[str] = Counter()
    reason_counter: Counter[str] = Counter()
    records: list[tuple[int, list[Any]]] = []
    for r in range(2, ws.max_row + 1):
        raw = get_cell(ws, r, headers, "대상 임대차계약 주소")
        final = source_addr(r)
        juso1 = norm(get_cell(ws, r, headers, "JUSO_판정"))
        juso2 = norm(get_cell(ws, r, headers, "JUSO_2차판정"))
        rec_jibun = get_cell(ws, r, headers, "JUSO_추천지번주소")
        base_jibun = get_cell(ws, r, headers, "JUSO_지번주소_1")
        rec_road = get_cell(ws, r, headers, "JUSO_추천도로명주소") or get_cell(
            ws, r, headers, "JUSO_도로명주소_1"
        )
        lot = parse_lot_addr(rec_jibun, base_jibun, raw, final)
        unit = unit_extract(raw, final, lot["addr"])
        bname = building_name(raw, lot["addr"], rec_road, final)
        prop = "집합건물" if unit["ho"] else "부동산구분 확인필요"
        phrase = norm(
            " ".join(
                x
                for x in [
                    lot["addr"],
                    bname,
                    suffix_dong(unit["bld_dong"]),
                    suffix_ho(unit["ho"]),
                ]
                if x
            )
        )
        reasons: list[str] = []
        if juso1 == "검색불가" or juso1.startswith("다중검출"):
            reasons.append(f"Juso 1차 {juso1}")
        if original_is_under_specified(raw):
            reasons.append(
                "원문 식별정보 부족: 지번/도로명건물번호/건물명 없이 법정동+호수만 있음"
            )
        if not lot["addr"]:
            reasons.append("지번주소 없음")
        if not unit["ho"]:
            reasons.append("호실 없음")
        unit_warning = unit_out_of_range(unit["bld_dong"], unit["ho"])
        if unit_warning:
            reasons.append(unit_warning)
        if juso2 in {"수동검토", "후보1_검토"}:
            reasons.append(f"Juso 2차 {juso2}")
        if has_extra_parcels(raw):
            reasons.append("외필지 포함: 대표지번 기준 조회 후 결과 대조")
        if unit["bld_dong"] and re.search(r"[가-힣]", unit["bld_dong"]):
            reasons.append("동표기 한글: A/B동 표기 병행검색")

        if not reasons:
            grade = "바로조회가능"
            strategy = "인터넷등기소 전체검색: 최종 검색용 주소 그대로 검색"
        elif (
            "지번주소 없음" in reasons
            or "호실 없음" in reasons
            or any(r.startswith("원문 식별정보 부족") for r in reasons)
        ):
            grade = "보완필요"
            strategy = "원문·경매기록·건축물대장 등으로 지번/호실 보완 필요"
        else:
            grade = "검토후조회"
            strategy = "전체검색어로 검색하되 결과의 건물명/동/호를 원문과 대조 후 발급"

        registry_values = [
            prop,
            lot["sido"],
            norm(" ".join(x for x in [lot["city"], lot["sigungu"]] if x)),
            norm(" ".join(x for x in [lot.get("eupmyeon", ""), lot["dong"]] if x)),
            lot["lot"],
            lot["addr"],
            bname,
            unit["bld_dong"],
            unit["floor"],
            unit["ho"],
            phrase,
            strategy,
            grade,
            "; ".join(reasons),
        ]
        set_row_values(ws, r, headers, REGISTRY_COLUMNS, registry_values)

        aux = []
        if lot["addr"] and unit["ho"]:
            aux.append(norm(f"{lot['addr']} {suffix_ho(unit['ho'])}"))
        if lot["addr"] and bname:
            aux.append(norm(f"{lot['addr']} {bname}"))
        aux = [x for i, x in enumerate(aux) if x and x not in aux[:i]]
        result = (
            "보완필요"
            if grade == "보완필요" or not (lot["addr"] and unit["ho"])
            else ("검토후조회" if grade == "검토후조회" else "바로조회가능")
        )
        whole_status = "보완필요" if result == "보완필요" else "전체검색가능"
        final_phrase = "" if result == "보완필요" else phrase
        final_aux = "" if result == "보완필요" else " / ".join(aux)
        set_row_values(
            ws, r, headers, WHOLE_COLUMNS, [final_phrase, final_aux, whole_status]
        )
        set_row_values(ws, r, headers, FINAL_COLUMNS, [final_phrase, result])
        final_counter[result] += 1
        for reason in reasons:
            reason_counter[reason] += 1
        records.append(
            (
                r,
                registry_values
                + [phrase, " / ".join(aux), whole_status, phrase, result],
            )
        )

    # 검토용 시트.
    review_cols = [
        "엑셀행",
        "원본행",
        "법무대리인",
        "대상 임대차계약 주소",
        "최종주소",
        "JUSO_판정",
        "JUSO_2차판정",
        "최종 검색용 주소",
        "주소검토결과",
        "등기_검토사유",
        "등기소_전체검색_보조검색어",
    ]
    for sheet_name, wanted in [
        ("등기소_바로조회가능", "바로조회가능"),
        ("등기소_검토필요", None),
    ]:
        sh = add_or_replace_sheet(wb, sheet_name)
        sh.append(review_cols)
        for r in range(2, ws.max_row + 1):
            result = get_cell(ws, r, headers, "주소검토결과")
            if (wanted and result != wanted) or (
                wanted is None and result == "바로조회가능"
            ):
                continue
            sh.append(
                [
                    r,
                    get_cell(ws, r, headers, "원본행"),
                    get_cell(ws, r, headers, "법무대리인"),
                    get_cell(ws, r, headers, "대상 임대차계약 주소"),
                    get_cell(ws, r, headers, "최종주소"),
                    get_cell(ws, r, headers, "JUSO_판정"),
                    get_cell(ws, r, headers, "JUSO_2차판정"),
                    get_cell(ws, r, headers, "최종 검색용 주소"),
                    result,
                    get_cell(ws, r, headers, "등기_검토사유"),
                    get_cell(ws, r, headers, "등기소_전체검색_보조검색어"),
                ]
            )
        sh.freeze_panes = "A2"
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(
                70, max(10, max(len(str(c.value or "")) for c in col) + 2)
            )

    summary_sheet = add_or_replace_sheet(wb, "최종검색주소_요약")
    summary_sheet.append(["구분", "건수"])
    for key_name in ["바로조회가능", "검토후조회", "보완필요"]:
        summary_sheet.append([key_name, final_counter.get(key_name, 0)])
    summary_sheet.append(["합계", sum(final_counter.values())])
    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 12

    ws.column_dimensions[
        ws.cell(1, headers["최종 검색용 주소"]).column_letter
    ].width = 70
    ws.column_dimensions[ws.cell(1, headers["주소검토결과"]).column_letter].width = 18

    output_file = (
        output_dir / f"{input_file.stem}_등기소전체검색용_최종검색주소추가.xlsx"
    )
    wb.save(output_file)
    summary = {
        "input": str(input_file),
        "output": str(output_file),
        "rows": ws.max_row - 1,
        "jusoFirstPass": dict(first_counter),
        "jusoSecondPassTargetRows": len(target_rows),
        "jusoSecondPass": dict(second_counter),
        "addressReview": dict(final_counter),
        "reviewReasons": dict(reason_counter),
    }
    summary_path = output_dir / f"{input_file.stem}_summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return summary
